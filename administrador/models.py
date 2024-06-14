from django.db import models
from django.contrib.auth.models import User
from cursos.models import Cursos
from estudiantes.models import Estudiante
from profesor.models import Profesor
from openpyxl import load_workbook
from django.shortcuts import get_object_or_404
from django.core.exceptions import ValidationError
from django.db import transaction

class ImportarCursosException(Exception):
    pass
# Create your models here.
class administrador(models.Model):
    user = models.OneToOneField(User, on_delete=models.CASCADE)
    codigo = models.PositiveBigIntegerField(null=True)


    def importar_cursos(self, file):
        wb = load_workbook(file)
        hoja = wb.active
        
        celda_nula = hoja['Z1'].value
        
        curso_codigo = hoja['A2'].value
        curso_nombre = hoja['C2'].value
        curso_periodo = hoja['F2'].value
        
        if(curso_codigo == celda_nula or curso_nombre == celda_nula or curso_periodo == celda_nula):
            raise ImportarCursosException("Hay campos de curso vacíos")
        
        profesor_identificacion = hoja['B3'].value
        profesor_nombre = hoja['C3'].value
        profesor_apellido = hoja['D3'].value
        profesor_email = hoja['E3'].value
        
        if(profesor_apellido == celda_nula or profesor_nombre == celda_nula or profesor_email == celda_nula or profesor_identificacion == celda_nula):
            raise ImportarCursosException("Hay campos de profesor vacios")
        
        username_p = profesor_nombre + str(profesor_identificacion)
        
        try:
            with transaction.atomic():
                try:
                    user = User.objects.get(email=profesor_email)
                    if (user.username != username_p or
                        user.first_name != profesor_nombre or
                        user.last_name != profesor_apellido):
                        raise ImportarCursosException(f"El correo electrónico de {profesor_nombre} ya está asociado a un usuario con credenciales diferentes.")
                except User.DoesNotExist:
                    user, creado = User.objects.get_or_create(username=username_p, email=profesor_email, first_name=profesor_nombre, last_name=profesor_apellido)
                    if creado:
                        dp = profesor_nombre[0] + str(profesor_identificacion) + profesor_apellido[0]
                        user.set_password(dp)
                        user.save()

                nuevo_profesor, profesor_creado = Profesor.objects.get_or_create(user=user, identificacion=profesor_identificacion)
                if profesor_creado:
                    nuevo_profesor.save()

                nuevo_curso, curso_creado = Cursos.objects.get_or_create(nombre=curso_nombre, codigo=curso_codigo, estado=True, periodoAcademico=curso_periodo, profesor=nuevo_profesor)
                if curso_creado:
                    nuevo_curso.save()

                try:
                    for fila in hoja.iter_rows(min_row=4, values_only=True):
                        estudiante_codigo = fila[0]
                        estudiante_nombre = fila[2]
                        estudiante_apellido = fila[3]
                        estudiante_email = fila[4]

                        if estudiante_codigo == "Fin":
                            return 'Importe realizado correctamente'

                        if estudiante_apellido == celda_nula or estudiante_codigo == celda_nula or estudiante_email == celda_nula or estudiante_nombre == celda_nula:
                            raise ImportarCursosException("Hay campos de estudiantes vacíos")

                        username_e = estudiante_nombre + str(estudiante_codigo)

                        try:
                            user = User.objects.get(email=estudiante_email)
                            if (user.username != username_e or
                                user.first_name != estudiante_nombre or
                                user.last_name != estudiante_apellido):
                                raise ImportarCursosException(f"El correo electrónico de {estudiante_nombre} {estudiante_apellido} ya está asociado a un usuario con credenciales diferentes.")
                        except User.DoesNotExist:
                            user, creado = User.objects.get_or_create(username=username_e, email=estudiante_email, first_name=estudiante_nombre, last_name=estudiante_apellido)
                            if creado:
                                dp = estudiante_nombre[0] + str(estudiante_codigo) + estudiante_apellido[0]
                                user.set_password(dp)
                                user.save()

                        nuevo_estudiante, creado = Estudiante.objects.get_or_create(user=user, codigo=estudiante_codigo)
                        if creado:
                            nuevo_estudiante.save()

                        nuevo_curso.estudiantes.add(nuevo_estudiante)
                except Exception as e:
                    # Si ocurre un error durante la importación de estudiantes, elimina el curso y el profesor creados
                    if curso_creado:
                        nuevo_curso.delete()
                    if profesor_creado:
                        nuevo_profesor.delete()
                    raise e
        except ImportarCursosException as e:
        # Manejo adicional de la excepción si es necesario
            raise e
                
    def importar_estudiantes(self, file):
        wb = load_workbook(file)
        hoja = wb.active
        celda_nula = hoja['Z1'].value
        
        curso_codigo = hoja['B1'].value
        
        curso = get_object_or_404(Cursos, codigo=curso_codigo)
        
        for fila in hoja.iter_rows(min_row=3, values_only=True):
            estudiante_codigo = fila[0]
            estudiante_nombre = fila[1]
            estudiante_apellido = fila[2]
            estudiante_email = fila[3]
            
            if(estudiante_codigo == "Fin"):
                return 'Importe realizado correctamente'
            
            if(estudiante_apellido == celda_nula or estudiante_codigo == celda_nula or estudiante_email == celda_nula or estudiante_nombre == celda_nula):
                raise ImportarCursosException("Hay campos de estudiantes vacios")

            username_e = estudiante_nombre + str(estudiante_codigo)
            
            try:
                user = User.objects.get(email=estudiante_email)
                if (user.username != username_e or
                        user.first_name != estudiante_nombre or
                        user.last_name != estudiante_apellido):
                    raise ImportarCursosException (f"El correo electrónico de {estudiante_nombre} {estudiante_apellido} ya está asociado a un usuario con credenciales diferentes.")
            except User.DoesNotExist:
                user, creado = User.objects.get_or_create(username=username_e, email=estudiante_email, first_name=estudiante_nombre, last_name=estudiante_apellido)
                if creado:
                    dp = estudiante_nombre[0] + str(estudiante_codigo) + estudiante_apellido[0]
                    user.set_password(dp)
                    user.save()

            nuevo_estudiante, creado = Estudiante.objects.get_or_create(user=user, codigo=estudiante_codigo)
            if creado:
                nuevo_estudiante.save()
            
            curso.estudiantes.add(nuevo_estudiante)
    
    def __str__(self):
        return self.user.username