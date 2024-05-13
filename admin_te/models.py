from django.db import models
from django.contrib.auth.models import User
from cursos.models import Cursos
from estudiantes.models import Estudiante
from profesor.models import Profesor
from openpyxl import load_workbook

class Admin_Te(models.Model):
    user = models.OneToOneField(User, on_delete=models.CASCADE)
    codigo = models.CharField(max_length=20)
    
    def read_file(self, file):
        wb = load_workbook(file)
        hoja = wb.active
        
        curso_codigo = hoja['A1'].value
        curso_nombre = hoja['C1'].value
        curso_periodo = hoja['E1'].value

        default_password = 'A123456B'
        
        profesor_identificacion = hoja['A2'].value
        profesor_nombre = hoja['C2'].value
        profesor_email = hoja['D2'].value
        user_profesor, creado = User.objects.get_or_create(username=profesor_nombre, email=profesor_email)
        if creado: 
            user_profesor.set_password(default_password)
            user_profesor.save()
        nuevo_profesor = Profesor.objects.create(user=user_profesor, identificacion=profesor_identificacion, telefono="123")
        nuevo_profesor.save()
        
        nuevo_curso = Cursos.objects.create(nombre=curso_nombre, codigo=curso_codigo, estado=True, periodoAcademico=curso_periodo, profesor=nuevo_profesor)
        nuevo_curso.save()
        
        for fila in hoja.iter_rows(min_row=3, values_only=True):
            estudiante_codigo = fila[0]
            estudiante_nombre = fila[2]
            estudiante_email = fila[3]
            user_estudiante, creado = User.objects.get_or_create(username=estudiante_nombre, email=estudiante_email)
            if creado:
                user_estudiante.set_password('A123456B')
                user_estudiante.save()
            nuevo_estudiante, creado = Estudiante.objects.get_or_create(user=user_estudiante, codigo=estudiante_codigo)
            nuevo_estudiante.save()
            
        def __str__(self):
            return self.user.username